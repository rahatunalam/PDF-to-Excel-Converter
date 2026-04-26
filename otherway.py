import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

pdf_path = "G:/Exam/Book_doc.pdf"

chapters = []
sections = []
hadiths = []

chapter_id = 1
section_id = 1
hadith_id = 1

current_chapter = None
current_section = None

# For hadith accumulation
current_hadith_text = ""
#This line of code creates a Regular Expression (regex) rule designed to find specific text patterns,
#likely Hadith numbers at the start of a line.
hadith_pattern = re.compile(r'^\[\s*[0-9০-৯]+\s*\]')

def is_bold(word):
    return "Bold" in word.get("fontname", "")

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        words = page.extract_words(extra_attrs=["fontname"])

        # Group words into lines
        lines = {}
        for word in words:
            y = round(word['top'],1)
            lines.setdefault(y, []).append(word)

        sorted_lines = sorted(lines.items(), key=lambda x: x[0])

        for _, line_words in sorted_lines:
            line_text = " ".join([w['text'] for w in line_words]).strip()

            if not line_text:
                continue

            # ---------------------------
            # Chapter detection (*)
            # ---------------------------
            if line_text.startswith("*"):
                if current_hadith_text:
                    hadiths.append({
                        "id": hadith_id,
                        "hadith": current_hadith_text.strip()
                    })
                    hadith_id += 1
                    current_hadith_text = ""

                current_chapter = line_text.replace("*", "").strip()
                chapters.append({
                    "id": chapter_id,
                    "name": current_chapter
                })
                chapter_id += 1
                continue

            # ---------------------------
            # Section detection (Bold)
            # ---------------------------
            bold_words = [w for w in line_words if is_bold(w)]

            # If at least 50% words are bold OR short line → treat as section
            if (len(bold_words) >= len(line_words) * 0.5) or (len(line_text) < 60):
                # Avoid capturing hadith lines
                if not hadith_pattern.match(line_text) and not line_text.startswith("*"):
                    current_section = line_text
                    sections.append({
                        "id": section_id,
                        "name": current_section
                    })
                    section_id += 1
                    continue

            # ---------------------------
            # Hadith detection
            # ---------------------------
            if hadith_pattern.match(line_text) or re.match(r'^\[[0-9০-৯]+\]$', line_text):
                if current_hadith_text:
                    hadiths.append({
                        "id": hadith_id,
                        "hadith": current_hadith_text.strip()
                    })
                    hadith_id += 1

                current_hadith_text = line_text
            else:
                if current_hadith_text:
                    current_hadith_text += " " + line_text

# Save last hadith
if current_hadith_text:
    hadiths.append({
        "id": hadith_id,
        "hadith": current_hadith_text.strip()
    })

# ---------------------------
# Create DataFrames
# ---------------------------
df_chapter = pd.DataFrame(chapters)
df_section = pd.DataFrame(sections)
df_hadith = pd.DataFrame(hadiths)

# ---------------------------
# Save to Excel
# ---------------------------
excel_path = "G:/Exam/output_full_hadith.xlsx"

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_chapter.to_excel(writer, sheet_name="chapter", index=False)
    df_section.to_excel(writer, sheet_name="section", index=False)
    df_hadith.to_excel(writer, sheet_name="hadith", index=False)

# ---------------------------
# Apply Kalpurush font
# ---------------------------
wb = load_workbook(excel_path)
font = Font(name="Kalpurush")

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font

wb.save(excel_path)

print("✅ Full multi-line hadith extraction completed!")