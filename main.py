import os
import fitz
import openpyxl
import re
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill



# ── Config ──────────────────────────────────────────────────────────────────
PDF_PATH  = "G:/Exam/Book_doc.pdf"
OUT_PATH  = "G:/Exam/Book_data1.xlsx"

BENGALI_DIGITS = '০১২৩৪৫৬৭৮৯'
HADITH_PATTERN = re.compile(r'^\[([' + BENGALI_DIGITS + r']+)\]')
# Pattern to strip leading [number] from hadith text
STRIP_NUM_PATTERN = re.compile(r'^\[[' + BENGALI_DIGITS + r']+\]\s*')

# ── Extraction ───────────────────────────────────────────────────────────────
def extract_all(pdf_path):
    doc = fitz.open(pdf_path)
    chapters, sections, hadiths = [], [], []
    seen_c, seen_s = set(), set()

    all_lines = []
    for page in doc:
        for block in page.get_text("dict")["blocks"]:
            if "lines" not in block:
                continue
            for line in block["lines"]:
                text = "".join(sp["text"] for sp in line["spans"]).strip()
                if text:
                    all_lines.append(text)
    doc.close()

    # Pass 1: chapters (*) and sections (**)
    for line in all_lines:
        if line.startswith("**"):
            name = line.lstrip("*").strip()
            if name and name not in seen_s:
                seen_s.add(name)
                sections.append(name)
        elif line.startswith("*"):
            name = (line.lstrip("*")
                    .replace("অধ্যায়:", "")
                    .replace("অধ্যায় :", "")
                    .strip())
            if name and name not in seen_c:
                seen_c.add(name)
                chapters.append(name)

    # Pass 2: hadiths (multi-line grouped by [bengali number])
    current = None
    for line in all_lines:
        if HADITH_PATTERN.match(line):
            if current is not None:
                hadiths.append(" ".join(current))
            # Remove the [bengali number] prefix before storing
            clean_line = STRIP_NUM_PATTERN.sub("", line).strip()
            current = [clean_line] if clean_line else []
        elif current is not None:
            if line.startswith("*"):
                hadiths.append(" ".join(current))
                current = None
            else:
                current.append(line)
    if current:
        hadiths.append(" ".join(current))

    return chapters, sections, hadiths

# ── Excel styling ────────────────────────────────────────────────────
FONT_NAME   = "Kalpurush"
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name=FONT_NAME, size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, title: str, headers: list, rows: list, col_widths: list):
    ws = wb.create_sheet(title)

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _border()

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font   = DATA_FONT
            cell.border = _border()
            cell.alignment = CENTER if col == 1 else LEFT
            if fill:
                cell.fill = fill

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 21

    ws.freeze_panes = "A2"
    return ws



def main():
    print("Extracting data from PDF …")
    chapters, sections, hadiths = extract_all(PDF_PATH)
    print(f"  Chapters : {len(chapters)}")
    print(f"  Sections : {len(sections)}")
    print(f"  Hadiths  : {len(hadiths)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "chapter",
                ["id", "name"],
                [(i, n) for i, n in enumerate(chapters, 1)],
                [8, 55])

    write_sheet(wb, "section",
                ["id", "name"],
                [(i, n) for i, n in enumerate(sections, 1)],
                [8, 65])

    write_sheet(wb, "hadith",
                ["id", "hadith"],
                [(i, t) for i, t in enumerate(hadiths, 1)],
                [8, 100])

    # Summary tab
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nSaved → {OUT_PATH}")

if __name__ == '__main__':
    main()